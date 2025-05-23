import datetime
from email.mime.multipart import MIMEMultipart
import smtplib
from email.mime.text import MIMEText
from typing import Dict, Optional
import os

from sqlalchemy.orm import Session

from models import AdminModifiedPresence, DailyPresence, User, HoursDefault
from serialization import DailyPresenceBase, HoursDefaultBase, UserBase
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from openpyxl.drawing.image import Image
from openpyxl.worksheet.page import PageMargins
from io import BytesIO
from datetime import time


def get_user_by_username(db: Session, work_email: str):
    return db.query(User).filter(User.work_email == work_email).first()


# Get Daily Presences for a specific user and month
def get_daily_presences(db: Session, user_id: int, year: int, month: int):
    return db.query(DailyPresence).filter(
        DailyPresence.user_id == user_id,
        DailyPresence.date.year == year,
        DailyPresence.date.month == month).all()


# Create or update a daily presence record
def create_or_update_daily_presence(db: Session, user_id: int, daily_presence_data: DailyPresenceBase):
    presence_record = db.query(DailyPresence).filter(
        DailyPresence.user_id == user_id,
        DailyPresence.date == daily_presence_data.date
    ).first()

    if presence_record:
        # Update existing record
        for key, value in daily_presence_data.dict().items():
            setattr(presence_record, key, value)
    else:
        # Create new record
        presence_record = DailyPresence(user_id=user_id, **daily_presence_data.dict())
        db.add(presence_record)

    db.commit()
    db.refresh(presence_record)
    return presence_record


def create_default_hours(db: Session, defaults: HoursDefaultBase):
    default_hours = db.query(HoursDefault).filter(HoursDefault.user_id == defaults.user_id,
                                                  HoursDefault.submitted_by_id == defaults.submitted_by_id).first()

    if not default_hours:
        default_hours = HoursDefault(**defaults.dict())
        db.add(default_hours)

    else:
        default_hours.entry_time_morning = defaults.entry_time_morning
        default_hours.exit_time_morning = defaults.exit_time_morning
        default_hours.entry_time_afternoon = defaults.entry_time_afternoon
        default_hours.exit_time_afternoon = defaults.exit_time_afternoon

    db.commit()
    db.refresh(default_hours)
    return default_hours


def get_user_default_hours(db: Session, user_id: int, submitted_by_id: int):
    default_hours = (db.query(HoursDefault).
                     filter(HoursDefault.user_id == user_id,
                            HoursDefault.submitted_by_id == submitted_by_id).
                     first())

    return default_hours


def get_user_by_id(db: Session, user_id: int):
    employee = db.query(User).filter(User.id == user_id).first()
    return employee


def get_hour_minute(timeFormat: int) -> (int, int):
    hoursPart = int(timeFormat) // 1
    minutesPart = (int(timeFormat) % 1) * 60
    return hoursPart, minutesPart


def has_submitted_presence(employee_id: int, year: str, month: str, db: Session) -> bool:
    presence = db.query(DailyPresence).filter_by(
        employee_id=employee_id,
        date=datetime.strptime(year+month, "%Y%m")
    ).first()
    return bool(presence)


def has_admin_submitted_presence(employee_id: int, year: str, month: str, db: Session) -> bool:
    presence = db.query(AdminModifiedPresence).filter_by(
        employee_id=employee_id,
        date=datetime.strptime(year+month, "%Y%m")
    ).first()
    return bool(presence)


def send_email_to_employee(receiver_email: str, subject: str, body: str):
    smtp_server = "smtps.aruba.it"
    smtp_port = 465
    smtp_user_sender = os.getenv("SMTP_USERNAME")
    smtp_sender_password = os.getenv("SMTP_PASSWORD")

    try:
        with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
            msg = MIMEMultipart("alternative")
            html_part = MIMEText(body, "html")
            msg.attach(html_part)
            server.login(smtp_user_sender, smtp_sender_password)
            msg.attach(html_part)
            msg["Subject"] = subject
            msg["From"] = smtp_user_sender
            msg["To"] = receiver_email
            server.sendmail(smtp_user_sender, receiver_email, msg.as_string())
    except Exception as e:
        print(f"Failed to send email to {receiver_email}: {e}")


def calculate_hours_per_day(morning_in, morning_out, afternoon_in, afternoon_out):
    total_hours = 0.0

    if morning_in and morning_out and morning_out > morning_in:
        total_hours += (datetime.combine(datetime.today(), morning_out) -
                        datetime.combine(datetime.today(), morning_in)).seconds / 3600

    if afternoon_in and afternoon_out and afternoon_out > afternoon_in:
        total_hours += (datetime.combine(datetime.today(), afternoon_out) -
                        datetime.combine(datetime.today(), afternoon_in)).seconds / 3600

    return total_hours


def create_excel_original(presence_data: [DailyPresence], employeeOverview: Dict=Optional):
    # Create Excel workbook
    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "Presenze mensile"
    if employeeOverview:
        sheet2 = workbook.create_sheet("Osservazione mensile")
    highlight_fill = PatternFill(start_color="0eff65", end_color="0eff65", fill_type="solid")
    weekend_holiday_fill = PatternFill(start_color="ffc7ab", end_color="ffc7ab", fill_type="solid")
    dayOff_fill = PatternFill(start_color="edec07", end_color="edec07", fill_type="solid")

    sheet1.append([
        "Date", "Entrata", "Uscita", "Entrata", "Uscita",
        "FERIE", "FESTIVITÀ NAZIONALE", "WEEKEND", "STRAORDINARIO", "PERMESSO", "NOTE"
    ])

    for i, day in enumerate(presence_data, start=2):
        sheet1.append([
            day.date,
            day.entry_time_morning,
            day.exit_time_morning,
            day.entry_time_afternoon,
            day.exit_time_afternoon,
            "Yes" if day.day_off else "No",
            "Yes" if day.national_holiday else "No",
            "Yes" if day.weekend else "No",
            day.extra_hours,
            day.time_off,
            day.notes
        ])

        sheet1.column_dimensions["A"].width = 12

        if day.weekend:
            for col in range(1, 12):
                sheet1.cell(row=i, column=col).fill = weekend_holiday_fill

        if day.day_off or day.national_holiday:
            for col in range(1, 12):
                sheet1.cell(row=i, column=col).fill = dayOff_fill

        if day.extra_hours != time(0,0):
            sheet1[f"I{i}"].fill = highlight_fill

        # Highlight time off if not "00:00"
        if day.time_off != time(0, 0):
            sheet1[f"J{i}"].fill = highlight_fill

        # Highlight notes if not empty
        if day.notes:
            sheet1[f"K{i}"].fill = highlight_fill

        if day.day_off:
            sheet1[f"F{i}"].fill = highlight_fill

    # Style the header row
    for cell in sheet1[1]:
        cell.font = Font(bold=True)
    if employeeOverview:
        sheet2.column_dimensions["A"].width = 26
        sheet2.append(["Metric", "Value"])
        sheet2.append(["Is Submitted", "Yes" if employeeOverview["isSubmitted"] else "No"])
        sheet2.append(["Total Worked Hours", employeeOverview["totalWorkedHoursInMonth"]])
        sheet2.append(["Total Extra Hours", employeeOverview["totalExtraHoursInMonth"]])
        sheet2.append(["Total Off Hours", employeeOverview["totalOffHoursInMonth"]])
        sheet2.append(["Total Off Days", employeeOverview["totalOffDaysInMonth"]])
        sheet2.append(["Total Expected Working Hours", employeeOverview["totalExpectedWorkingHours"]])
        sheet2.append(["Notes", employeeOverview["notes"]])

    for cell in sheet2[1]:
        cell.font = Font(bold=True)

    excel_output = BytesIO()
    workbook.save(excel_output)
    excel_output.seek(0)

    return excel_output


def create_excel_modified(presence_data: [DailyPresence], employee: UserBase):

    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "Presenze mensile"
    sheet1.page_setup.fitToWidth = 1
    sheet1.page_setup.fitToHeight = 0  
    sheet1.page_margins = PageMargins(top=0.5, bottom=0.5, left=0.3, right=0.3)

    
    weekend_holiday_fill = PatternFill(start_color="bcbcbc", end_color="bcbcbc", fill_type="solid")
    festivita_fill = PatternFill(start_color="ffada7", end_color="ffada7", fill_type="solid")
    bold_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    current_dir = os.getcwd()
    img = Image(os.path.join(current_dir, 'logo.png'))
    img.anchor = 'A1'
    sheet1.add_image(img)
    for _ in range(6):
        sheet1.append([])

    sheet1.append(["Cognome:", employee.surname, "Nome", employee.name])
    sheet1.append([])
    sheet1.append([])
    sheet1.append(["Anno", presence_data[0].date.year, "Mese", presence_data[0].date.month])
    sheet1.append([])
    sheet1.append([])

    sheet1.append(["Date", "Entrata", "Uscita", "Entrata", "Uscita", "Ore", "Note"])

    for day in presence_data:
        Notes = ''
        if day.modified_day_off:
            Notes += "FERIE " 
        if day.modified_national_holiday:
            Notes += "FESTIVITÀ NAZIONALE "  
        if day.modified_extra_hours.isoformat() != '00:00:00':
            Notes += f"STRAORDINARIO:{day.modified_extra_hours.strftime('%H:%M')} "  
        if day.modified_time_off.isoformat() != '00:00:00':
            Notes += f"PERMESSO:{day.modified_time_off.strftime('%H:%M')} "
        if day.modified_illness:
            Notes += f"Malattia:{day.modified_illness} "
        if day.modified_notes:
            Notes += day.modified_notes

        def format_time(t: time) -> str:
            return "" if t.strftime("%H:%M") == "00:00" else t.strftime("%H:%M")

        workedHoursInDay = calculate_hours_per_day(day.modified_entry_time_morning,
                                                   day.modified_exit_time_morning,
                                                   day.modified_entry_time_afternoon,
                                                   day.modified_exit_time_afternoon)

        row_data = [
            day.date.day,  
            format_time(day.modified_entry_time_morning),
            format_time(day.modified_exit_time_morning),
            format_time(day.modified_entry_time_afternoon),
            format_time(day.modified_exit_time_afternoon),
            workedHoursInDay,
            Notes
        ]
        sheet1.append(row_data)

        current_row = sheet1.max_row
        for col in range(1, 8):
            cell = sheet1.cell(row=current_row, column=col)
            cell.alignment = center_alignment
            cell.border = thin_border

        if day.modified_weekend or day.modified_day_off :
            for col in range(1, 8):
                sheet1.cell(row=current_row, column=col).fill = weekend_holiday_fill
        if day.modified_national_holiday:
            for col in range(1, 8):
                sheet1.cell(row=current_row, column=col).fill = festivita_fill

    for row in [7, 10]:
        for col in range(1, 4):
            cell = sheet1.cell(row=row, column=col)
            cell.alignment = center_alignment
            if col in [1, 3]:
                cell.font = bold_font

    for col in range(1, 8):
        cell = sheet1.cell(row=13, column=col)
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border

    end_row = sheet1.max_row
    for row in range(13, end_row + 1):
        for col in range(1, 8):
            sheet1.cell(row=row, column=col).border = thin_border

    sheet1.column_dimensions["A"].width = 10
    sheet1.column_dimensions["B"].width = 8
    sheet1.column_dimensions["C"].width = 8
    sheet1.column_dimensions["D"].width = 8
    sheet1.column_dimensions["E"].width = 8
    sheet1.column_dimensions["F"].width = 6
    sheet1.column_dimensions["G"].width = 52

    excel_output = BytesIO()
    workbook.save(excel_output)
    excel_output.seek(0)
    return excel_output